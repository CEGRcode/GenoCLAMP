import os
import json
from collections import namedtuple, defaultdict
import numpy as np
import xml.dom.minidom as dom
import openpyxl
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
try:
    from cairosvg import svg2png
    use_cairosvg = True
except ImportError:
    use_cairosvg = False
from io import BytesIO
from utils import trim_motif

def write_aligned_transfac(cluster, filename):
    with open(filename, 'w') as f:
        for i in range(len(cluster.items)):
            motif_id = cluster.items[i].source[0]
            f.write('AC\t{}\n'.format(motif_id))
            f.write('XX\n')
            f.write('ID\t{}\n'.format(motif_id))
            f.write('P0\tA\tC\tG\tT\n')
            pfm = cluster.aligned_pfms[i, :, :]
            for j in range(pfm.shape[0]):
                f.write('{:02d}\t{:06f}\t{:06f}\t{:06f}\t{:06f}\n'.format(j + 1, *pfm[j, :]))
            f.write('XX\n//\n')
            
def write_consensus_transfac(cluster, filename, info_thresh=.5):
    c = cluster.idx
    trimmed_pfm, _, _, _ = trim_motif(cluster.aligned_pfms, info_thresh)
    with open(filename, 'w') as f:
        f.write('AC\t{}\n'.format('cluster{}'.format(c)))
        f.write('XX\n')
        f.write('ID\t{}\n'.format('cluster{}'.format(c)))
        f.write('P0\tA\tC\tG\tT\n')
        for j in range(trimmed_pfm.shape[0]):
            f.write('{:02d}\t{:06f}\t{:06f}\t{:06f}\t{:06f}\n'.format(j + 1, *trimmed_pfm[j, :]))
        f.write('XX\n//\n')

def write_consensus_meme(cluster, filename, info_thresh=.5):
    c = cluster.idx
    trimmed_pfm, _, _, _ = trim_motif(cluster.aligned_pfms, info_thresh)
    with open(filename, 'w') as f:
        f.write('MEME version 5\n')
        f.write('ALPHABET= ACGT\n')
        f.write('MOTIF {0} {0}\n'.format('cluster{}'.format(c)))
        f.write('letter-probability matrix: alength= 4 w= {} nsites= {} E= 0\n'.format(
            trimmed_pfm.shape[0], sum(item.source[1] for item in cluster.items)))
        for j in range(trimmed_pfm.shape[0]):
            f.write('{:06f} {:06f} {:06f} {:06f}\n'.format(*(trimmed_pfm[j, :] / np.sum(trimmed_pfm[j, :]))))

clamp_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
with open('{}/logo_symbols/glyphs.json'.format(clamp_dir), 'r') as f:
    glyph_data = json.load(f)
with open('{}/logo_symbols/symbol_library.json'.format(clamp_dir), 'r') as f:
    symbol_library = json.load(f)

ColoredGlyph = namedtuple('ColoredGlyph', ['path', 'color'])

class DNASymbol:
    max_bits = symbol_library['DNA']['max_bits']
    DNA_alphabet = tuple(ColoredGlyph(path=glyph_data[s['name']]['path'], color=s['color']) \
                         for s in symbol_library['DNA']['symbols'])

    @classmethod
    def get_symbol(cls, i):
        return cls.DNA_alphabet[i]

class RNASymbol:
    max_bits = symbol_library['RNA']['max_bits']
    RNA_alphabet = tuple(ColoredGlyph(path=glyph_data[s['name']]['path'], color=s['color']) \
                         for s in symbol_library['RNA']['symbols'])
    
    @classmethod
    def get_symbol(cls, i):
        return cls.RNA_alphabet[i]

class AASymbol:
    max_bits = symbol_library['AA']['max_bits']
    AA_alphabet = tuple(ColoredGlyph(path=glyph_data[s['name']]['path'], color=s['color']) \
                        for s in symbol_library['AA']['symbols'])
    
    @classmethod
    def get_symbol(cls, i):
        return cls.AA_alphabet[i]

def plot_logo_stack(aligned_pfms, symbol=DNASymbol, glyph_width=100, stack_height=200):
    '''
    Plots a stack of logos for a stack of aligned PFMs and outputs it as an SVG
    To save the SVG:
        doc = plot_logo_stack(aligned_pfms)
        with open(filename, 'w') as f:
            doc.writexml(f, addindent='\t', newl='\n')
    '''
    height, width, _ = aligned_pfms.shape

    document = dom.Document()
    svg = document.appendChild(document.createElement('svg'))
    svg.setAttribute('baseProfile', 'full')
    svg.setAttribute('version', '1.1')
    svg.setAttribute('xmlns', 'http://www.w3.org/2000/svg')
    svg.setAttribute('viewBox', '0 0 {} {}'.format(width * glyph_width, height * stack_height))

    logo_stack = svg.appendChild(document.createElement('g'))

    for y, pfm in enumerate(aligned_pfms):
        row = logo_stack.appendChild(document.createElement('g'))
        row.setAttribute('transform', 'translate(0 {})'.format(y * stack_height))
        for i, pwv in enumerate(pfm):
            n = np.sum(pwv)
            if n == 0:
                continue
            vec = pwv / n
            vec[vec > 0] *= np.log2(vec[vec > 0])
            bits = sum(vec, symbol.max_bits)
            heights = pwv / n * bits / symbol.max_bits * stack_height
            idx = np.argsort(pwv)
    
            stack = row.appendChild(document.createElement('g'))
            stack.setAttribute('transform', 'translate({} 0)'.format(i * glyph_width))
    
            y_offset = 0
            for j in idx:
                if heights[j] == 0:
                    continue
                base = symbol.get_symbol(j)
                y_offset += heights[j]
                glyph = stack.appendChild(document.createElement('path'))
                glyph.setAttribute('d', base.path)
                glyph.setAttribute('fill', base.color)
                glyph.setAttribute('transform', 'matrix({} 0 0 {} 0 {})'.format(glyph_width / 100., heights[j] / 100., stack_height - y_offset))

    return svg

def write_bed_file(cluster, filename):
    with open(filename, 'w') as f:
        for (chrom, start, end, strand), sources in cluster.sites.items():
            f.write('{}\t{}\t{}\t{}\t.\t{}\n'.format(chrom, start, end, ';'.join(source[0] for source in sources), strand))

def idx2col(idx):
    col = []
    while idx:
        idx, rem = divmod(idx - 1, 26)
        col.append(chr(ord('A') + rem))
    return ''.join(reversed(col))

nucleotide_map = np.array(['N', 'A', 'C', 'M', 'G', 'R', 'S', 'V', 'T', 'W', 'Y', 'H', 'K', 'D', 'B', 'N'])
bitmap = 1 << np.arange(4)
def pfm2iupac(pfm):
    mask = pfm > np.max(pfm, axis=1, keepdims=True) / 2
    mask[np.sum(pfm * mask, axis=1) < .6 * np.sum(pfm, axis=1), :] = 0
    return ''.join(nucleotide_map[np.sum(mask * bitmap, axis=1)])

fonts = defaultdict(lambda: InlineFont(rFont='Courier', color='888888'))
fonts.update(A=InlineFont(rFont='Courier', color='FF0000'), C=InlineFont(rFont='Courier', color='0000FF'),
             G=InlineFont(rFont='Courier', color='FFA500'), T=InlineFont(rFont='Courier', color='228B22'))

def write_summary_excel(engine, maximal_clusters, filename, info_thresh=1., sites=False):
    wb = openpyxl.Workbook()
    ws1 = wb.worksheets[0]
    ws1.title = 'Motif clusters'
    ws1['A1'] = 'Cluster ID'
    ws1['B1'] = 'Consensus logo'
    ws1.column_dimensions['B'].width = 55
    ws1['C1'] = 'IUPAC sequence'
    ws1.column_dimensions['C'].width = 32
    ws1['D1'] = 'BLLR'
    ws1['E1'] = 'Scaled BLLR'
    ws1['F1'] = 'N component motifs'
    ws1['G1'] = 'Sources'
    if sites:
        ws1['H1'] = 'N sites'
        full_sources = set()
        for c in maximal_clusters:
            for item in engine.clusters[c].items:
                full_sources.add(item.source[0])
        source_col_map = {source: idx2col(i + 8) for i, source in enumerate(full_sources)}
        for source, col in source_col_map.items():
            ws1['{}1'.format(col)] = '{} loci'.format(source)

        ws2 = wb.create_sheet('Reference points')
        ws2['A1'] = 'Chrom'
        ws2['B1'] = 'Start'
        ws2['C1'] = 'End'
        ws2['D1'] = 'Cluster ID'
        ws2['E1'] = 'Strand'
        ws2['F1'] = 'Sources'

        total_sites = 0

    for i, c in enumerate(sorted(maximal_clusters, key=lambda c_: engine.clusters[c_].llr, reverse=True)):
        ws1.row_dimensions[i + 2].height = 25
        pfm, left_offset, right_offset, trimmed = trim_motif(engine.clusters[c].aligned_pfms, info_thresh=info_thresh)
        iupac_seq = CellRichText()
        for nuc in pfm2iupac(pfm):
            iupac_seq.append(TextBlock(fonts[nuc], nuc))
        logo = plot_logo_stack(np.expand_dims(pfm, 0))

        ws1['A{}'.format(i + 2)] = 'cluster{}'.format(c)
        if use_cairosvg:
            ws1.add_image(openpyxl.drawing.image.Image(BytesIO(svg2png(logo.toxml(), dpi=200, output_height=30))), anchor='B{}'.format(i + 2))
        else:
            cell = 'B{}'.format(i + 2)
            link = 'cluster{0}/cluster{0}_consensus-motif.svg'.format(c)
            ws1[cell] = link
            ws1[cell].hyperlink = link
            ws1[cell].style = 'Hyperlink'
        ws1['C{}'.format(i + 2)] = iupac_seq
        ws1['D{}'.format(i + 2)] = engine.clusters[c].llr
        ws1['D{}'.format(i + 2)].number_format = '0.00'
        ws1['E{}'.format(i + 2)] = engine.clusters[c].llr / len(engine.clusters[c].items)
        ws1['E{}'.format(i + 2)].number_format = '0.00'
        ws1['F{}'.format(i + 2)] = len(engine.clusters[c].items)

        sources = set()
        for item in engine.clusters[c].items:
            sources.add(item.source[0])
        ws1['G{}'.format(i + 2)] = ';'.join(sources)

        if sites:
            ws1['H{}'.format(i + 2)] = len(engine.clusters[c].sites)
            site_counter = {}
            for site, source_set in engine.clusters[c].sites.items():
                for source in source_set:
                    site_counter.setdefault(source, 0)
                    site_counter[source] += 1
            for source, nsites in site_counter.items():
                ws1['{}{}'.format(source_col_map[source], i + 2)] = nsites
                ws1['{}{}'.format(source_col_map[source], i + 2)].fill = PatternFill(start_color='FFFFF261', end_color='FFFFF261', fill_type='solid')
        
            for (chrom, start, end, strand), source_set in engine.clusters[c].sites.items():
                if strand == '+':
                    start += left_offset
                    end -= right_offset
                else:
                    start += right_offset
                    end -= left_offset
                ws2['A{}'.format(total_sites + 2)] = chrom
                ws2['B{}'.format(total_sites + 2)] = start
                ws2['C{}'.format(total_sites + 2)] = end
                ws2['D{}'.format(total_sites + 2)] = 'cluster{}'.format(i + 1)
                ws2['E{}'.format(total_sites + 2)] = strand
                ws2['F{}'.format(total_sites + 2)] = ';'.join(source_set)
                total_sites += 1
        
    ws3 = wb.create_sheet('Input motifs')
    ws3['A1'] = 'Motif ID'
    ws3['B1'] = 'Logo'
    ws3.column_dimensions['B'].width = 55
    ws3['C1'] = 'IUPAC sequence'
    ws3.column_dimensions['C'].width = 32
    ws3['D1'] = 'Source'
    ws3['E1'] = 'N sites'
    ws3['F1'] = 'E-value'
    
    for i, item in enumerate(engine.items):
        pfm = item.pfm
        doc = plot_logo_stack(np.expand_dims(pfm, 0))
        iupac_seq = CellRichText()
        for nuc in pfm2iupac(pfm):
            iupac_seq.append(TextBlock(fonts[nuc], nuc))
        
        ws3.row_dimensions[i + 2].height = 25
    
        ws3['A{}'.format(i + 2)] = 'input{}'.format(i + 1)
        if use_cairosvg:
            ws3.add_image(openpyxl.drawing.image.Image(BytesIO(svg2png(doc.toxml(), dpi=200, output_height=30))), anchor='B{}'.format(i + 2))
        else:
            ws3['B{}'.format(i + 2)] = 'Install cairosvg to view input logo'
        ws3['C{}'.format(i + 2)] = iupac_seq
        ws3['D{}'.format(i + 2)] = item.source[0]
        ws3['E{}'.format(i + 2)] = item.source[1]
        ws3['F{}'.format(i + 2)] = item.source[2]
    
    wb.save(filename)